import openpyxl
from configExcel import file_path


# Función para leer los datos desde el archivo Excel
def Data_excel(file_path):
    wb = openpyxl.load_workbook(file_path)  # Cargar el archivo Excel
    sheet = wb.active  # Seleccionar la hoja activa

    data_dict = {}  # Inicializar un diccionario para almacenar los datos

    print(
        f"Nombre de la hoja activa: {sheet.title}"
    )  # Imprimir el nombre de la hoja activa

    for row in range(
        2, sheet.max_row + 1
    ):  # Iterar desde la fila 2 hasta la última fila
        # Leer todos los valores de la fila en una lista
        row_data = [
            sheet.cell(row=row, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]

        # Almacenar los datos en el diccionario usando el número de fila como clave
        data_dict[row] = row_data

        # Imprimir los datos leídos para verificar
        print(f"Fila {row}: {row_data}")

    return data_dict


# Función para consultar los datos con base en el índice de la fila y la posición


def print_data_for_row(row_index, data_dict):
    if row_index not in data_dict:
        print(f"No se encontraron datos para la fila {row_index}")
        return

    # Extraer los valores de la lista correspondiente a la fila
    row_data = data_dict[row_index]
    Musica = row_data[0]
    print(f"{Musica}")

    # Mostrar los datos consultados
    print(f"Datos para la fila {row_index}:")
    for index, value in enumerate(row_data):
        print(f"Columna {index + 1}: {value}")


file_path_data = file_path

# Leer los datos del Excel
data = Data_excel(file_path_data)

# Cambia el índice para consultar otras filas
# Debes usar el índice correcto para la fila que tiene datos
print_data_for_row(2, data)  # Cambié el índice a 2, porque la fila 2 tiene datos.
